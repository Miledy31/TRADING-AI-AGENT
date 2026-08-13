import os, math, time, requests
from datetime import datetime, timezone
from openpyxl import load_workbook

XLSX = "Trading_Dashboard_V1_1.xlsx"
ALPHA_KEY = os.getenv("ALPHAVANTAGE_API_KEY", "")
COINGECKO_KEY = os.getenv("COINGECKO_API_KEY", "")

STOCKS = ["AAPL","NVDA","MSFT","SPY","QQQ","IWM"]
CRYPTO = {"BTC":"bitcoin","ETH":"ethereum","SOL":"solana"}

def sma(values, n):
    return sum(values[:n]) / n if len(values) >= n else None

def rsi(closes, period=14):
    if len(closes) <= period: return None
    gains, losses = [], []
    for i in range(1, period+1):
        d = closes[i-1] - closes[i]
        gains.append(max(d,0))
        losses.append(max(-d,0))
    ag, al = sum(gains)/period, sum(losses)/period
    if al == 0: return 100.0
    return 100 - (100/(1 + ag/al))

def score_setup(price, s20, s50, rsi14, change_pct, volume_ratio=None):
    score = 0
    reasons = []
    if s20 and price > s20: score += 20; reasons.append("price above SMA20")
    if s50 and price > s50: score += 20; reasons.append("price above SMA50")
    if s20 and s50 and s20 > s50: score += 20; reasons.append("SMA20 above SMA50")
    if rsi14 is not None and 50 <= rsi14 <= 70: score += 15; reasons.append("RSI in bullish zone")
    if change_pct is not None and change_pct > 0: score += 10; reasons.append("positive momentum")
    if volume_ratio is not None and volume_ratio >= 1.2: score += 15; reasons.append("volume expansion")
    status = "POTENTIAL" if score >= 70 else "WATCH" if score >= 50 else "AVOID"
    return score, min(score,100), status, "; ".join(reasons)

def alpha_daily(symbol):
    if not ALPHA_KEY:
        raise RuntimeError("Set ALPHAVANTAGE_API_KEY first.")
    url="https://www.alphavantage.co/query"
    p={"function":"TIME_SERIES_DAILY","symbol":symbol,"outputsize":"compact","apikey":ALPHA_KEY}
    j=requests.get(url,params=p,timeout=30).json()
    ts=j.get("Time Series (Daily)")
    if not ts: raise RuntimeError(f"{symbol}: no daily data returned: {j}")
    rows=sorted(ts.items(), reverse=True)
    closes=[float(x[1]["4. close"]) for x in rows]
    vols=[float(x[1]["5. volume"]) for x in rows]
    price=closes[0]
    prev=closes[1] if len(closes)>1 else price
    change=(price/prev-1)*100
    s20=sma(closes,20); s50=sma(closes,50)
    r=rsi(closes)
    vr=(vols[0]/(sum(vols[1:21])/20)) if len(vols)>=21 else None
    return price,change,r,s20,s50,vr

def cg_markets():
    if COINGECKO_KEY:
        url="https://api.coingecko.com/api/v3/coins/markets"
        headers={"x-cg-demo-api-key":COINGECKO_KEY}
    else:
        url="https://api.coingecko.com/api/v3/coins/markets"
        headers={}
    p={"vs_currency":"usd","ids":",".join(CRYPTO.values()),"price_change_percentage":"24h"}
    j=requests.get(url,params=p,headers=headers,timeout=30).json()
    return {x["id"]: x for x in j}

def update():
    wb=load_workbook(XLSX)
    dash=wb["Dashboard"]; sc=wb["Scanner"]
    now=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    scanner_row=2
    live_row=17

    for sym in STOCKS:
        try:
            price,chg,r,s20,s50,vr=alpha_daily(sym)
            score,conf,status,reason=score_setup(price,s20,s50,r,chg,vr)
            vals=[sym,"Stock","Day/Swing","1D",price,chg,r,s20,s50,"",vr,"Trend/Momentum","LONG",price,price*0.98,price*1.04,score/25,score,conf,status,reason,now]
            for c,v in enumerate(vals,1): sc.cell(scanner_row,c,v)
            live=[sym,"Stock","Day/Swing",price,chg,r,s20,s50,vr,"Trend/Momentum",conf,status,now]
            for c,v in enumerate(live,1): dash.cell(live_row,c,v)
            scanner_row+=1; live_row+=1
        except Exception as e:
            print("Stock error:",e)

    try:
        markets=cg_markets()
        for sym,cgid in CRYPTO.items():
            x=markets[cgid]; price=float(x["current_price"]); chg=float(x.get("price_change_percentage_24h") or 0)
            score=max(0,min(100,50+chg*2))
            status="POTENTIAL" if score>=70 else "WATCH" if score>=50 else "AVOID"
            reason="24h momentum from CoinGecko market data"
            vals=[sym,"Crypto","Day/Swing","1D",price,chg,"","","","",x.get("total_volume"),"Momentum","LONG",price,price*0.95,price*1.10,2.0,score,score,status,reason,now]
            for c,v in enumerate(vals,1): sc.cell(scanner_row,c,v)
            live=[sym,"Crypto","Day/Swing",price,chg,"","","",x.get("total_volume"),"Momentum",score,status,now]
            for c,v in enumerate(live,1): dash.cell(live_row,c,v)
            scanner_row+=1; live_row+=1
    except Exception as e:
        print("Crypto error:",e)

    dash["E4"]="CONNECTED"
    dash["E5"]=now
    wb.save(XLSX)
    print("Updated", now)

if __name__=="__main__":
    update()
